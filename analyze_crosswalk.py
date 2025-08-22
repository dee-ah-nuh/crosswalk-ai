#!/usr/bin/env python3
"""
Analyze the crosswalk Excel template to understand structure and formulas
"""

import pandas as pd
import openpyxl
from openpyxl.formula import Tokenizer
import json

def analyze_crosswalk_template(file_path):
    """Analyze the crosswalk Excel template"""
    print(f"Analyzing crosswalk template: {file_path}")
    
    # Load with openpyxl to get formulas
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # Print all sheet names
    print(f"Available sheets: {wb.sheetnames}")
    
    # Focus on Medical sheet if it exists
    if 'Medical' in wb.sheetnames:
        sheet = wb['Medical']
        print(f"\n=== MEDICAL SHEET ANALYSIS ===")
        print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        
        # Get column headers and formulas in top rows
        print(f"\n=== COLUMN STRUCTURE ===")
        headers = []
        formulas = []
        
        for col in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=1, column=col)
            formula_cells = []
            
            # Check first few rows for formulas
            for row in range(1, min(6, sheet.max_row + 1)):  # Check first 5 rows
                cell = sheet.cell(row=row, column=col)
                if cell.value and str(cell.value).startswith('='):
                    formula_cells.append({
                        'row': row,
                        'formula': str(cell.value),
                        'display_value': cell.displayed_value if hasattr(cell, 'displayed_value') else None
                    })
            
            headers.append({
                'column': col,
                'header': str(header_cell.value) if header_cell.value else f"Column_{col}",
                'formulas': formula_cells
            })
        
        # Display column info
        for h in headers[:15]:  # Show first 15 columns to avoid too much output
            print(f"Column {h['column']}: {h['header']}")
            if h['formulas']:
                for f in h['formulas']:
                    print(f"  Row {f['row']}: {f['formula']}")
            print()
    
    # Also load with pandas to see data
    try:
        print(f"\n=== PANDAS DATA PREVIEW ===")
        df = pd.read_excel(file_path, sheet_name='Medical' if 'Medical' in wb.sheetnames else 0, nrows=10)
        print(f"Data shape: {df.shape}")
        print("\nColumn names:")
        for i, col in enumerate(df.columns):
            print(f"{i+1}: {col}")
        
        print(f"\nFirst few rows:")
        print(df.head(3).to_string())
        
    except Exception as e:
        print(f"Error reading with pandas: {e}")

if __name__ == "__main__":
    analyze_crosswalk_template("attached_assets/CLIENT_GDP_Crosswalk_1755826997501.xlsx")